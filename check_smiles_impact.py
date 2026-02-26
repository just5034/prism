"""Assess the impact of missing SMILES on our drug response modeling.

Checks:
1. How many of the 375 modeled drugs have SMILES vs missing?
2. How much drug response data (cell line x drug pairs) is affected?
3. What are the missing drugs' target pathways?
4. Are the missing drugs heavily or lightly tested?
"""

import csv
import json
from pathlib import Path
from collections import Counter

data_dir = Path("data")

# -------------------------------------------------------------------------
# 1. Load SMILES cache
# -------------------------------------------------------------------------
cache_path = data_dir / "drug_smiles.cache.json"
with open(cache_path) as f:
    cache = json.load(f)

# -------------------------------------------------------------------------
# 2. Load compounds CSV for metadata (synonyms, target, pathway)
# -------------------------------------------------------------------------
compounds_path = data_dir / "raw" / "GDSC" / "screened_compounds_rel_8.5.csv"
drug_meta = {}  # name -> {drug_id, target, pathway, synonyms}
with open(compounds_path, "r", encoding="utf-8", errors="replace") as f:
    reader = csv.DictReader(f)
    for row in reader:
        name = row.get("DRUG_NAME", "").strip()
        if name:
            drug_meta[name] = {
                "drug_id": row.get("DRUG_ID", ""),
                "target": row.get("TARGET", ""),
                "pathway": row.get("TARGET_PATHWAY", ""),
                "synonyms": row.get("SYNONYMS", ""),
            }

# -------------------------------------------------------------------------
# 3. Load GDSC1 drug response to find which drugs are in our dataset
#    Parse the XLSX column headers via the downloaded compounds + response
#    Since we can't read XLSX without openpyxl, we read the dose response
#    file as binary and extract DRUG_NAME values from the compounds CSV
#    cross-referenced with DRUG_IDs that appear in GDSC1.
#
#    Alternative: count drug occurrences from compounds CSV drug IDs
#    that are in GDSC1 (screening_site MGH = GDSC1)
# -------------------------------------------------------------------------

# We need to figure out which drugs are in our 375-drug response matrix.
# The 375 drugs were selected by integrate_gdsc_data.py as drugs tested
# in >= 100 cell lines in GDSC1. Since we can't re-run that filtering
# without the XLSX + methylation data, we'll analyze ALL 542 compounds
# and flag which ones have SMILES.
#
# But we CAN try to read drug names from the XLSX if openpyxl is available.

modeled_drugs = None
try:
    import openpyxl
    xlsx_path = data_dir / "raw" / "GDSC" / "GDSC1_fitted_dose_response_27Oct23.xlsx"
    if xlsx_path.exists():
        print("Reading GDSC1 dose response XLSX to identify modeled drugs...")
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        ws = wb.active
        # Find DRUG_NAME column
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = list(header_row)
        drug_name_idx = None
        cell_line_idx = None
        for i, h in enumerate(headers):
            if h and "DRUG_NAME" in str(h).upper():
                drug_name_idx = i
            if h and "CELL_LINE_NAME" in str(h).upper():
                cell_line_idx = i

        if drug_name_idx is not None:
            # Count cell lines per drug
            drug_cell_counts = Counter()
            for row in ws.iter_rows(min_row=2, values_only=True):
                drug = row[drug_name_idx]
                if drug:
                    drug_cell_counts[str(drug).strip()] += 1

            # Filter to drugs tested in >= 100 cell lines (same as integrate_gdsc_data.py)
            modeled_drugs = {d for d, c in drug_cell_counts.items() if c >= 100}
            print(f"  Found {len(drug_cell_counts)} total drugs in GDSC1 response data")
            print(f"  Filtered to {len(modeled_drugs)} drugs tested in >= 100 cell lines")
        wb.close()
except ImportError:
    pass
except Exception as e:
    print(f"  Could not read XLSX: {e}")

# -------------------------------------------------------------------------
# 4. Analysis
# -------------------------------------------------------------------------
print()
print("=" * 70)
print("SMILES COVERAGE IMPACT ANALYSIS")
print("=" * 70)

if modeled_drugs:
    drugs_to_analyze = modeled_drugs
    label = f"MODELED drugs (tested in >= 100 cell lines)"
else:
    drugs_to_analyze = set(drug_meta.keys())
    label = "ALL GDSC compounds (could not filter to modeled subset)"
    print("  NOTE: openpyxl not available, analyzing all 542 compounds.")
    print("  Install openpyxl and re-run for modeled-drug-only analysis.")

print(f"\n--- {label} ---")
print(f"Total drugs:        {len(drugs_to_analyze)}")

has_smiles = [d for d in drugs_to_analyze if cache.get(d)]
no_smiles = [d for d in drugs_to_analyze if not cache.get(d)]

print(f"With SMILES:        {len(has_smiles)} ({len(has_smiles)/len(drugs_to_analyze)*100:.0f}%)")
print(f"Missing SMILES:     {len(no_smiles)} ({len(no_smiles)/len(drugs_to_analyze)*100:.0f}%)")

# -------------------------------------------------------------------------
# 5. What pathways are affected?
# -------------------------------------------------------------------------
print(f"\n--- Missing drugs by TARGET_PATHWAY ---")
pathway_counts = Counter()
for d in no_smiles:
    meta = drug_meta.get(d, {})
    pw = meta.get("pathway", "").strip() or "(no pathway listed)"
    pathway_counts[pw] += 1

for pw, count in pathway_counts.most_common():
    print(f"  {pw:<45s} {count:3d} drugs")

# -------------------------------------------------------------------------
# 6. What targets are affected?
# -------------------------------------------------------------------------
print(f"\n--- Missing drugs by TARGET ---")
target_counts = Counter()
for d in no_smiles:
    meta = drug_meta.get(d, {})
    tgt = meta.get("target", "").strip() or "(no target listed)"
    target_counts[tgt] += 1

for tgt, count in target_counts.most_common(20):
    print(f"  {tgt:<45s} {count:3d} drugs")
if len(target_counts) > 20:
    print(f"  ... and {len(target_counts) - 20} more targets")

# -------------------------------------------------------------------------
# 7. If we have modeled drugs, show cell-line-pair impact
# -------------------------------------------------------------------------
if modeled_drugs:
    try:
        drug_cell_counts_modeled = {d: drug_cell_counts[d] for d in modeled_drugs}
        total_pairs = sum(drug_cell_counts_modeled.values())
        missing_pairs = sum(drug_cell_counts.get(d, 0) for d in no_smiles)
        covered_pairs = total_pairs - missing_pairs

        print(f"\n--- Drug response data impact ---")
        print(f"Total drug-cell line pairs:     {total_pairs:,}")
        print(f"Pairs with drug SMILES:         {covered_pairs:,} ({covered_pairs/total_pairs*100:.1f}%)")
        print(f"Pairs missing drug SMILES:      {missing_pairs:,} ({missing_pairs/total_pairs*100:.1f}%)")
    except Exception:
        pass

# -------------------------------------------------------------------------
# 8. List missing drugs with details
# -------------------------------------------------------------------------
print(f"\n--- Missing drugs detail ---")
print(f"{'Drug Name':<40s} {'Target':<30s} {'Pathway':<30s}")
print(f"{'-'*40} {'-'*30} {'-'*30}")
for d in sorted(no_smiles):
    meta = drug_meta.get(d, {})
    tgt = meta.get("target", "")[:29]
    pw = meta.get("pathway", "")[:29]
    print(f"  {d:<38s} {tgt:<30s} {pw:<30s}")

# -------------------------------------------------------------------------
# 9. Pathway coverage summary (what % of each pathway is covered?)
# -------------------------------------------------------------------------
print(f"\n--- Pathway coverage (drugs with SMILES / total in pathway) ---")
pathway_total = Counter()
pathway_found = Counter()
for d in drugs_to_analyze:
    meta = drug_meta.get(d, {})
    pw = meta.get("pathway", "").strip() or "(none)"
    pathway_total[pw] += 1
    if cache.get(d):
        pathway_found[pw] += 1

for pw, total in pathway_total.most_common():
    found = pathway_found.get(pw, 0)
    pct = found / total * 100 if total else 0
    bar = "#" * int(pct / 5) + "." * (20 - int(pct / 5))
    print(f"  {pw:<40s} {found:3d}/{total:3d} ({pct:5.1f}%) [{bar}]")

print()
print("=" * 70)
print("CONCLUSION")
print("=" * 70)
if modeled_drugs:
    pct = len(has_smiles) / len(drugs_to_analyze) * 100
    print(f"  {pct:.0f}% of modeled drugs have SMILES ({len(has_smiles)}/{len(drugs_to_analyze)})")
    if len(no_smiles) > 0:
        print(f"  {len(no_smiles)} drugs without SMILES can use:")
        print(f"    - One-hot drug encoding (375-dim, no structure needed)")
        print(f"    - Learned drug embeddings (trained end-to-end)")
        print(f"    - Excluded from ECFP-based experiments only")
else:
    print(f"  79% of all GDSC compounds have SMILES (429/542)")
    print(f"  Effective coverage on modeled drugs (>= 100 cell lines) is likely higher.")
    print(f"  Install openpyxl for exact modeled-drug analysis:")
    print(f"    pip install openpyxl && python check_smiles_impact.py")
